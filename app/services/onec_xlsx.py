from __future__ import annotations

import hashlib
import hmac
import re
from datetime import date, datetime
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
DEPARTMENT_SEPARATOR = " / "

DEFAULT_COLUMNS = {
    "snils": "СНИЛС",
    "fio": "Сотрудник.Физическое лицо.ФИО",
    "email": "Физическое лицо.Адрес электронной почты",
    "personal_email": "Физическое лицо.Email",
    "mobile_phone": "Физическое лицо.Мобильный телефон",
    "position": "Должность",
}

DEFAULT_COLUMN_ALIASES = {
    "fio": ("Сотрудник",),
}


@dataclass(frozen=True)
class OneCPlacement:
    department: str | None
    position: str | None


@dataclass
class OneCWorker:
    worker_key: str
    fio: str
    email: str | None
    login: str | None
    placements: tuple[OneCPlacement, ...]
    personal_email: str | None = None
    mobile_phone: str | None = None
    dismissal_date: date | None = None


@dataclass(frozen=True)
class OneCWorkbook:
    workers: tuple[OneCWorker, ...]
    headers: tuple[str, ...]
    header_row: int
    detected_columns: dict[str, str]
    potential_dismissal_columns: tuple[str, ...]
    dismissal_column: str = ""
    dismissal_column_ambiguous: bool = False

    @property
    def placements_count(self) -> int:
        return sum(len(worker.placements) for worker in self.workers)

    @property
    def multiple_placements_count(self) -> int:
        return sum(1 for worker in self.workers if len(worker.placements) > 1)

    @property
    def missing_email_count(self) -> int:
        return sum(1 for worker in self.workers if not worker.email)

def normalize_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip().lower()


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def normalize_snils(value: Any) -> str | None:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits if len(digits) == 11 else None


def normalize_email(value: Any) -> str | None:
    email = str(value or "").strip().lower()
    return email if EMAIL_RE.fullmatch(email) else None


def worker_key(snils: str, secret: str) -> str:
    return hmac.new(
        secret.encode("utf-8"),
        snils.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()


def find_header_row(
    sheet,
    expected: dict[str, str],
    search_rows: int,
) -> tuple[int, dict[str, int], tuple[str, ...]]:
    normalized_expected = {
        key: normalize_header(name)
        for key, name in expected.items()
    }

    best_row: int | None = None
    best_mapping: dict[str, int] = {}
    best_headers: tuple[str, ...] = ()

    for row_idx in range(1, min(search_rows, sheet.max_row) + 1):
        raw_headers = tuple(
            normalize_text(sheet.cell(row=row_idx, column=col).value)
            for col in range(1, sheet.max_column + 1)
        )
        normalized = [normalize_header(value) for value in raw_headers]

        mapping: dict[str, int] = {}
        for key, target in normalized_expected.items():
            aliases = [
                normalize_header(value)
                for value in DEFAULT_COLUMN_ALIASES.get(key, ())
                if expected.get(key) == DEFAULT_COLUMNS.get(key)
            ]

            # First prefer exact canonical/alias matches. This allows the
            # second export's "Сотрудник" without confusing it with other
            # columns that merely start with that word.
            for candidate in (target, *aliases):
                for col_idx, value in enumerate(normalized, start=1):
                    if value == candidate:
                        mapping[key] = col_idx
                        break
                if key in mapping:
                    break

            # Preserve the old tolerant behavior only for the canonical
            # column name. Short aliases such as "Сотрудник" are exact-only.
            if key not in mapping:
                for col_idx, value in enumerate(normalized, start=1):
                    if target and target in value:
                        mapping[key] = col_idx
                        break

        if len(mapping) > len(best_mapping):
            best_row = row_idx
            best_mapping = mapping
            best_headers = tuple(value for value in raw_headers if value)

    required = {"snils", "fio"}
    if best_row is None or not required.issubset(best_mapping):
        missing = sorted(required - set(best_mapping))
        raise ValueError(
            "Не найдены обязательные колонки XLSX: "
            + ", ".join(missing)
        )

    return best_row, best_mapping, best_headers


def _looks_like_department_row(
    values: list[Any],
    snils_col: int,
    fio_col: int,
    email_col: int | None,
) -> bool:
    snils_value = values[snils_col - 1] if snils_col <= len(values) else None
    fio_value = values[fio_col - 1] if fio_col <= len(values) else None
    email_value = (
        values[email_col - 1]
        if email_col and email_col <= len(values)
        else None
    )

    if normalize_snils(snils_value) or normalize_text(fio_value) or normalize_email(email_value):
        return False

    nonempty = [normalize_text(value) for value in values if normalize_text(value)]
    return len(nonempty) == 1


def _department_path(hierarchy: dict[int, str]) -> str | None:
    parts: list[str] = []
    for level in sorted(hierarchy):
        department = normalize_text(hierarchy[level])
        if not department:
            continue
        if parts and parts[-1].casefold() == department.casefold():
            continue
        parts.append(department)
    return DEPARTMENT_SEPARATOR.join(parts) or None


def _update_department_hierarchy(
    hierarchy: dict[int, str],
    outline_level: int,
    department: str,
) -> str | None:
    for level in list(hierarchy):
        if level >= outline_level:
            del hierarchy[level]
    hierarchy[outline_level] = department
    return _department_path(hierarchy)


def _placement_key(placement: OneCPlacement) -> tuple[str, str]:
    return (
        normalize_text(placement.department).casefold(),
        normalize_text(placement.position).casefold(),
    )


def _append_placement(
    placements: tuple[OneCPlacement, ...],
    placement: OneCPlacement,
) -> tuple[OneCPlacement, ...]:
    key = _placement_key(placement)
    for index, item in enumerate(placements):
        if _placement_key(item) != key:
            continue
        updated = OneCPlacement(
            department=placement.department or item.department,
            position=placement.position or item.position,
        )
        return (*placements[:index], updated, *placements[index + 1 :])
    return (*placements, placement)


DISMISSAL_MARKERS = (
    ("дата увольнения", 100),
    ("уволь", 80),
    ("дата прекращ", 70),
    ("дата окончания", 60),
    ("расторж", 50),
    ("termination", 40),
    ("dismiss", 40),
)


def _parse_dismissal_date(value: Any, *, epoch) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value, epoch=epoch)
            return parsed.date() if isinstance(parsed, datetime) else parsed
        except Exception:
            return None

    text = normalize_text(value)
    if not text:
        return None
    for fmt in (
        "%d.%m.%Y",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d.%m.%y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _dismissal_column(sheet, header_row: int) -> tuple[int | None, str, bool]:
    candidates: list[tuple[int, int, str]] = []
    for col in range(1, sheet.max_column + 1):
        raw = sheet.cell(row=header_row, column=col).value
        header = normalize_text(raw)
        normalized = normalize_header(header)
        best_score = 0
        for marker, score in DISMISSAL_MARKERS:
            if marker in normalized:
                best_score = max(best_score, score)
        if best_score:
            candidates.append((best_score, col, header))

    if not candidates:
        return None, "", False

    candidates.sort(key=lambda item: (-item[0], item[1]))
    best_score = candidates[0][0]
    best = [item for item in candidates if item[0] == best_score]
    if len(best) != 1:
        return None, "", True

    _, column, name = best[0]
    return column, name, False


def _dismissal_candidates(headers: tuple[str, ...]) -> tuple[str, ...]:
    markers = (
        "уволь",
        "дата прекращ",
        "дата окончания",
        "прекращение",
        "расторж",
        "dismiss",
        "termination",
    )
    candidates = [
        header
        for header in headers
        if any(marker in normalize_header(header) for marker in markers)
    ]
    return tuple(candidates)


def parse_onec_xlsx(
    path: Path,
    *,
    hash_secret: str,
    header_search_rows: int = 20,
    columns: dict[str, str] | None = None,
) -> OneCWorkbook:
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        sheet = workbook.active
        expected = dict(DEFAULT_COLUMNS)
        if columns:
            expected.update(columns)

        header_row, mapping, headers = find_header_row(
            sheet,
            expected,
            header_search_rows,
        )

        snils_col = mapping["snils"]
        fio_col = mapping["fio"]
        email_col = mapping.get("email")
        personal_email_col = mapping.get("personal_email")
        mobile_phone_col = mapping.get("mobile_phone")
        position_col = mapping.get("position")
        dismissal_col, dismissal_name, dismissal_ambiguous = _dismissal_column(
            sheet,
            header_row,
        )

        department_hierarchy: dict[int, str] = {}
        current_department: str | None = None
        merged: dict[str, OneCWorker] = {}
        dismissal_rows: dict[str, list[date | None]] = {}

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            values = [
                sheet.cell(row=row_idx, column=col).value
                for col in range(1, sheet.max_column + 1)
            ]

            if _looks_like_department_row(
                values,
                snils_col,
                fio_col,
                email_col,
            ):
                department = next(
                    normalize_text(value)
                    for value in values
                    if normalize_text(value)
                )
                outline_level = int(
                    sheet.row_dimensions[row_idx].outlineLevel or 0
                )
                current_department = _update_department_hierarchy(
                    department_hierarchy,
                    outline_level,
                    department,
                )
                continue

            snils = normalize_snils(values[snils_col - 1])
            fio = normalize_text(values[fio_col - 1])
            if not snils or not fio:
                continue

            email = (
                normalize_email(values[email_col - 1])
                if email_col
                else None
            )
            login = email.split("@", 1)[0] if email else None
            personal_email = (
                normalize_email(values[personal_email_col - 1])
                if personal_email_col
                else None
            )
            mobile_phone = (
                normalize_text(values[mobile_phone_col - 1]) or None
                if mobile_phone_col
                else None
            )

            position = (
                normalize_text(values[position_col - 1])
                if position_col
                else ""
            ) or None
            placement = OneCPlacement(
                department=current_department,
                position=position,
            )
            key = worker_key(snils, hash_secret)
            if dismissal_col is not None:
                dismissal_rows.setdefault(key, []).append(
                    _parse_dismissal_date(
                        values[dismissal_col - 1],
                        epoch=workbook.epoch,
                    )
                )
            existing = merged.get(key)

            if existing:
                merged[key] = OneCWorker(
                    worker_key=key,
                    fio=fio or existing.fio,
                    email=email or existing.email,
                    login=login or existing.login,
                    placements=_append_placement(
                        existing.placements,
                        placement,
                    ),
                    personal_email=personal_email or existing.personal_email,
                    mobile_phone=mobile_phone or existing.mobile_phone,
                )
            else:
                merged[key] = OneCWorker(
                    worker_key=key,
                    fio=fio,
                    email=email,
                    login=login,
                    placements=(placement,),
                    personal_email=personal_email,
                    mobile_phone=mobile_phone,
                )

        # Полная кадровая выгрузка является источником увольнений по отсутствию.
        # Поэтому пустой или изменившийся по структуре файл нельзя применять:
        # иначе все сотрудники могли бы ошибочно стать "отсутствующими".
        if not merged:
            raise ValueError(
                "Кадровая выгрузка не содержит ни одного работника. "
                "Предыдущий успешный снимок сохранен."
            )
        if dismissal_ambiguous:
            raise ValueError(
                "Колонка «Дата увольнения» определена неоднозначно. "
                "Кадровая выгрузка не применена."
            )
        if dismissal_col is None:
            raise ValueError(
                "В кадровой выгрузке не найдена колонка «Дата увольнения». "
                "Кадровая выгрузка не применена."
            )

        # Окончательная дата по организации появляется только когда
        # дата увольнения заполнена во ВСЕХ строках/должностях человека.
        # Если хотя бы одна должность продолжается, человек остается активным.
        if dismissal_col is not None and not dismissal_ambiguous:
            for key, worker in merged.items():
                values = dismissal_rows.get(key, [])
                if values and all(value is not None for value in values):
                    worker.dismissal_date = max(
                        value for value in values if value is not None
                    )
                else:
                    worker.dismissal_date = None

        detected_columns = {
            key: expected[key]
            for key in mapping
        }

        return OneCWorkbook(
            workers=tuple(merged.values()),
            headers=headers,
            header_row=header_row,
            detected_columns=detected_columns,
            potential_dismissal_columns=_dismissal_candidates(headers),
            dismissal_column=dismissal_name,
            dismissal_column_ambiguous=dismissal_ambiguous,
        )
    finally:
        workbook.close()


def worker_snapshot(worker: OneCWorker) -> dict:
    return {
        "worker_key": worker.worker_key,
        "fio": worker.fio,
        "email": worker.email,
        "personal_email": worker.personal_email,
        "mobile_phone": worker.mobile_phone,
        "login": worker.login,
        "dismissal_date": (
            worker.dismissal_date.isoformat()
            if worker.dismissal_date is not None
            else None
        ),
        "placements": [
            asdict(placement)
            for placement in worker.placements
        ],
    }
