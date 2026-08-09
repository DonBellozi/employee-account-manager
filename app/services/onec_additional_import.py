from __future__ import annotations

import hashlib
import json
import os
import re
import tempfile
import threading
from datetime import date, datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import desc, select
from sqlalchemy.orm import Session

from app.config import Settings
from app.models import HRPerson, HRSourceRecord, OneCImportRun
from app.models_onec_sources import HREmploymentState, OneCAdditionalSource
from app.services.hr_registry_multisource import MultiSourceHRRegistryViewService
from app.services.onec_imap import OneCAttachment, OneCImapService
from app.services.onec_xlsx import (
    DEFAULT_COLUMNS,
    find_header_row,
    normalize_header,
    normalize_snils,
    parse_onec_xlsx,
    worker_key,
    worker_snapshot,
)


SUCCESSFUL_STATUSES = {"success", "partial", "duplicate"}

DISMISSAL_MARKERS = (
    ("дата увольнения", 100),
    ("уволь", 80),
    ("дата прекращ", 70),
    ("дата окончания", 60),
    ("расторж", 50),
    ("termination", 40),
    ("dismiss", 40),
)


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _safe_source_dir(source_id: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9._-]+", "_", source_id.strip().lower())
    return value or "source"


def _parse_date(value, *, epoch) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value, epoch=epoch)
            return parsed.date() if isinstance(parsed, datetime) else parsed
        except Exception:
            return None

    text = " ".join(str(value).split()).strip()
    if not text:
        return None
    for fmt in (
        "%d.%m.%Y",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d.%m.%y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


class OneCAdditionalImportService:
    """Импорт одного дополнительного кадрового источника."""

    _lock = threading.Lock()

    def __init__(
        self,
        settings: Settings,
        db: Session,
        source: OneCAdditionalSource,
    ):
        self.settings = settings
        self.db = db
        self.source = source
        self.base_dir = (
            Path(settings.onec_data_dir)
            / "sources"
            / _safe_source_dir(source.source_id)
        )
        self.archive_dir = self.base_dir / "archive"
        self.current_file = self.base_dir / "current.xlsx"
        self.snapshot_file = self.base_dir / "current_snapshot.json"
        self.report_file = self.base_dir / "last_analysis.json"

    @property
    def hash_secret(self) -> str:
        return (
            self.settings.onec_worker_hash_secret.strip()
            or self.settings.app_secret_key
        )

    def find_latest(self) -> OneCAttachment:
        return OneCImapService(self.settings).find_latest_attachment(
            folder=self.source.imap_folder,
            sender_filter=self.source.sender_filter,
            attachment_filename=self.source.attachment_filename,
        )

    def analyze_latest(self, *, trigger: str = "manual") -> dict:
        if trigger not in {"manual", "scheduled", "startup"}:
            raise ValueError("Неизвестный тип запуска импорта 1С")
        if not self._lock.acquire(blocking=False):
            raise RuntimeError("Импорт дополнительной выгрузки уже выполняется")

        run = OneCImportRun(
            trigger=trigger,
            status="running",
            source_id=self.source.source_id,
            started_at=utcnow(),
        )
        self.db.add(run)
        self.db.commit()
        self.db.refresh(run)

        incoming_path: Path | None = None
        try:
            self.base_dir.mkdir(parents=True, exist_ok=True)
            self.archive_dir.mkdir(parents=True, exist_ok=True)

            attachment = self.find_latest()
            self._fill_mail(run, attachment)

            duplicate = self.db.scalars(
                select(OneCImportRun)
                .where(
                    OneCImportRun.source_id == self.source.source_id,
                    OneCImportRun.file_hash == attachment.file_hash,
                    OneCImportRun.status.in_(SUCCESSFUL_STATUSES),
                    OneCImportRun.id != run.id,
                )
                .order_by(desc(OneCImportRun.id))
                .limit(1)
            ).first()

            if duplicate is not None:
                active_records = list(
                    self.db.scalars(
                        select(HRSourceRecord).where(
                            HRSourceRecord.source_id
                            == self.source.source_id,
                            HRSourceRecord.is_present.is_(True),
                        )
                    ).all()
                )
                reconciliation = MultiSourceHRRegistryViewService(
                    self.settings,
                    self.db,
                ).reconcile_all()

                run.status = "duplicate"
                run.message = (
                    f"Файл уже обработан в импорте № {duplicate.id}. "
                    "Сверка AD/Zimbra обновлена."
                )
                run.workers_count = duplicate.workers_count
                run.placements_count = duplicate.placements_count
                run.completed_at = utcnow()
                self.db.commit()
                return {
                    "source": self.source.name,
                    "source_id": self.source.source_id,
                    "status": "duplicate",
                    "workers_count": len(active_records),
                    "missing_email_count": sum(
                        1
                        for record in active_records
                        if not record.corporate_email.strip()
                    ),
                    "registry": reconciliation,
                    "message": run.message,
                }

            incoming_path = self._write_incoming(attachment.payload)
            workbook = parse_onec_xlsx(
                incoming_path,
                hash_secret=self.hash_secret,
                header_search_rows=self.settings.onec_header_search_rows,
            )
            dismissal = self._dismissal_dates(incoming_path)
            current_snapshot = {
                worker.worker_key: {
                    **worker_snapshot(worker),
                    "dismissal_date": (
                        dismissal["dates"].get(worker.worker_key).isoformat()
                        if dismissal["dates"].get(worker.worker_key)
                        else None
                    ),
                }
                for worker in workbook.workers
            }
            previous_snapshot = self._load_snapshot()
            comparison = self._compare(previous_snapshot, current_snapshot)

            sync = self._sync_registry(
                workbook=workbook,
                dismissal_dates=dismissal["dates"],
            )
            reconciliation = MultiSourceHRRegistryViewService(
                self.settings,
                self.db,
            ).reconcile_all()

            report = {
                "analyzed_at": datetime.now().replace(microsecond=0).isoformat(),
                "source": self.source.name,
                "source_id": self.source.source_id,
                "mail_domain": self.source.mail_domain,
                "mail": {
                    "uid": attachment.uid,
                    "message_date": attachment.message_date,
                    "sender": attachment.sender,
                    "subject": attachment.subject,
                    "filename": attachment.filename,
                    "file_hash": attachment.file_hash,
                },
                "workers_count": len(workbook.workers),
                "placements_count": workbook.placements_count,
                "missing_email_count": workbook.missing_email_count,
                "comparison": comparison,
                "dismissal_column": dismissal["column"],
                "dismissal_column_ambiguous": dismissal["ambiguous"],
                "dismissal_dates_count": sum(
                    1 for value in dismissal["dates"].values() if value
                ),
                "employment": sync,
                "registry": reconciliation,
            }

            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            archive_filename = (
                f"{stamp}_{attachment.file_hash[:12]}_"
                f"{Path(attachment.filename).name}"
            )
            self._atomic_write_bytes(
                self.archive_dir / archive_filename,
                attachment.payload,
            )
            self._atomic_write_bytes(self.current_file, attachment.payload)
            self._atomic_write_json(self.snapshot_file, current_snapshot)
            self._atomic_write_json(self.report_file, report)

            reconcile_errors = reconciliation.get("errors") or []
            run.status = "partial" if reconcile_errors else "success"
            run.archive_filename = archive_filename
            run.workers_count = len(workbook.workers)
            run.placements_count = workbook.placements_count
            run.new_workers = comparison["new_workers"]
            run.missing_workers = comparison["missing_workers"]
            run.changed_workers = comparison["changed_workers"]
            run.message = (
                "Выгрузка обработана, сверка частично завершилась с ошибкой."
                if reconcile_errors
                else "Выгрузка обработана и сверена."
            )
            run.completed_at = utcnow()
            self.db.commit()
            return {
                **report,
                "status": run.status,
                "run_id": run.id,
                "message": run.message,
            }

        except Exception as exc:
            self.db.rollback()
            failed = self.db.get(OneCImportRun, run.id)
            if failed is not None:
                failed.status = "failed"
                failed.error_message = str(exc)[:4000]
                failed.message = "Импорт не выполнен."
                failed.completed_at = utcnow()
                self.db.commit()
            raise
        finally:
            if incoming_path is not None:
                incoming_path.unlink(missing_ok=True)
            self._lock.release()

    def _fill_mail(
        self,
        run: OneCImportRun,
        attachment: OneCAttachment,
    ) -> None:
        run.mail_uid = attachment.uid
        run.message_date = attachment.message_date
        run.sender = attachment.sender
        run.subject = attachment.subject
        run.filename = attachment.filename
        run.file_hash = attachment.file_hash
        self.db.commit()

    def _dismissal_dates(self, path: Path) -> dict:
        workbook = load_workbook(path, data_only=True, read_only=False)
        try:
            sheet = workbook.active
            header_row, mapping, _ = find_header_row(
                sheet,
                DEFAULT_COLUMNS,
                self.settings.onec_header_search_rows,
            )
            snils_col = mapping["snils"]

            candidates: list[tuple[int, int, str]] = []
            for col in range(1, sheet.max_column + 1):
                raw = sheet.cell(row=header_row, column=col).value
                header = " ".join(str(raw or "").replace("\n", " ").split()).strip()
                normalized = normalize_header(header)
                best_score = 0
                for marker, score in DISMISSAL_MARKERS:
                    if marker in normalized:
                        best_score = max(best_score, score)
                if best_score:
                    candidates.append((best_score, col, header))

            if not candidates:
                return {"column": "", "ambiguous": False, "dates": {}}

            candidates.sort(key=lambda item: (-item[0], item[1]))
            best_score = candidates[0][0]
            best = [item for item in candidates if item[0] == best_score]
            if len(best) != 1:
                return {
                    "column": "",
                    "ambiguous": True,
                    "dates": {},
                }

            _, dismissal_col, dismissal_name = best[0]
            result: dict[str, date | None] = {}
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                snils = normalize_snils(
                    sheet.cell(row=row_idx, column=snils_col).value
                )
                if not snils:
                    continue
                key = worker_key(snils, self.hash_secret)
                parsed = _parse_date(
                    sheet.cell(row=row_idx, column=dismissal_col).value,
                    epoch=workbook.epoch,
                )
                if key not in result or (
                    parsed is not None
                    and (result[key] is None or parsed > result[key])
                ):
                    result[key] = parsed

            return {
                "column": dismissal_name,
                "ambiguous": False,
                "dates": result,
            }
        finally:
            workbook.close()

    def _sync_registry(
        self,
        *,
        workbook,
        dismissal_dates: dict[str, date | None],
    ) -> dict[str, int]:
        now = utcnow()
        today = datetime.now(ZoneInfo(self.settings.app_timezone)).date()
        source_id = self.source.source_id
        current_keys = {worker.worker_key for worker in workbook.workers}

        records = self.db.scalars(
            select(HRSourceRecord).where(
                HRSourceRecord.source_id == source_id
            )
        ).all()
        by_worker = {row.worker_key: row for row in records}

        employment_rows = self.db.scalars(
            select(HREmploymentState).where(
                HREmploymentState.source_id == source_id
            )
        ).all()
        employment_by_worker = {
            row.worker_key: row for row in employment_rows
        }

        people = (
            self.db.scalars(
                select(HRPerson).where(
                    HRPerson.worker_key.in_(current_keys)
                )
            ).all()
            if current_keys
            else []
        )
        people_by_worker = {row.worker_key: row for row in people}

        created_people = 0
        created_records = 0
        dismissed = 0
        scheduled = 0
        active = 0

        for worker in workbook.workers:
            person = people_by_worker.get(worker.worker_key)
            if person is None:
                person = HRPerson(
                    worker_key=worker.worker_key,
                    fio=worker.fio,
                    first_seen_at=now,
                    last_seen_at=now,
                )
                self.db.add(person)
                people_by_worker[worker.worker_key] = person
                created_people += 1
            else:
                person.fio = worker.fio
                person.last_seen_at = now

            placements_json = json.dumps(
                [
                    {
                        "department": placement.department or "",
                        "position": placement.position or "",
                    }
                    for placement in worker.placements
                ],
                ensure_ascii=False,
                sort_keys=True,
            )

            record = by_worker.get(worker.worker_key)
            if record is None:
                record = HRSourceRecord(
                    worker_key=worker.worker_key,
                    source_id=source_id,
                    source_name=self.source.name,
                    fio=worker.fio,
                    corporate_email=worker.email or "",
                    personal_email=worker.personal_email or "",
                    login=worker.login or "",
                    placements_json=placements_json,
                    is_present=True,
                    first_seen_at=now,
                    last_seen_at=now,
                )
                self.db.add(record)
                by_worker[worker.worker_key] = record
                created_records += 1
            else:
                record.source_name = self.source.name
                record.fio = worker.fio
                record.corporate_email = worker.email or ""
                record.personal_email = worker.personal_email or ""
                record.login = worker.login or ""
                record.placements_json = placements_json
                record.is_present = True
                record.last_seen_at = now

            if worker.email:
                if record.zimbra_status == "no_email":
                    record.zimbra_status = "not_checked"
                if worker.login and record.ad_status == "no_login":
                    record.ad_status = "not_checked"
                if record.reconciliation_status == "issue" and not record.reconciliation_error:
                    record.reconciliation_status = "not_checked"
                    record.reconciled_at = None
            else:
                record.zimbra_status = "no_email"
                if not worker.login:
                    record.ad_status = "no_login"
                record.reconciliation_status = "issue"
                record.reconciliation_error = ""
                record.reconciled_at = now

            dismissal_date = dismissal_dates.get(worker.worker_key)
            if dismissal_date is None:
                status = "active"
                reason = "current_export"
                active += 1
            elif dismissal_date > today:
                status = "scheduled"
                reason = "dismissal_date"
                scheduled += 1
            else:
                status = "dismissed"
                reason = "dismissal_date"
                dismissed += 1

            employment = employment_by_worker.get(worker.worker_key)
            if employment is None:
                employment = HREmploymentState(
                    worker_key=worker.worker_key,
                    source_id=source_id,
                    source_name=self.source.name,
                    fio=worker.fio,
                    first_seen_at=now,
                )
                self.db.add(employment)
                employment_by_worker[worker.worker_key] = employment

            employment.source_name = self.source.name
            employment.fio = worker.fio
            employment.status = status
            employment.is_present = True
            employment.dismissal_date = dismissal_date
            employment.status_reason = reason
            employment.last_seen_at = now
            employment.updated_at = now

        absent = 0
        for worker_key, record in by_worker.items():
            if worker_key in current_keys:
                continue
            if record.is_present:
                record.is_present = False
            employment = employment_by_worker.get(worker_key)
            if employment is None:
                employment = HREmploymentState(
                    worker_key=worker_key,
                    source_id=source_id,
                    source_name=self.source.name,
                    fio=record.fio,
                    first_seen_at=record.first_seen_at,
                )
                self.db.add(employment)
                employment_by_worker[worker_key] = employment
            employment.source_name = self.source.name
            employment.fio = record.fio
            employment.status = "dismissed"
            employment.is_present = False
            employment.status_reason = "absent_from_export"
            employment.updated_at = now
            absent += 1

        self.db.commit()
        return {
            "created_people": created_people,
            "created_source_records": created_records,
            "active": active,
            "scheduled": scheduled,
            "dismissed_by_date": dismissed,
            "dismissed_by_absence": absent,
        }

    def _load_snapshot(self) -> dict[str, dict]:
        if not self.snapshot_file.is_file():
            return {}
        try:
            value = json.loads(
                self.snapshot_file.read_text(encoding="utf-8")
            )
        except (OSError, json.JSONDecodeError):
            return {}
        return value if isinstance(value, dict) else {}

    @staticmethod
    def _meaningful(value: dict) -> dict:
        return {
            "fio": value.get("fio") or "",
            "email": value.get("email") or "",
            "login": value.get("login") or "",
            "dismissal_date": value.get("dismissal_date"),
            "placements": value.get("placements") or [],
        }

    @classmethod
    def _compare(
        cls,
        previous: dict[str, dict],
        current: dict[str, dict],
    ) -> dict[str, int]:
        previous_keys = set(previous)
        current_keys = set(current)
        changed = {
            key
            for key in previous_keys & current_keys
            if cls._meaningful(previous.get(key) or {})
            != cls._meaningful(current.get(key) or {})
        }
        return {
            "new_workers": len(current_keys - previous_keys),
            "missing_workers": len(previous_keys - current_keys),
            "changed_workers": len(changed),
        }

    def _write_incoming(self, payload: bytes) -> Path:
        handle = tempfile.NamedTemporaryFile(
            prefix=".incoming_",
            suffix=".xlsx",
            dir=self.base_dir,
            delete=False,
        )
        try:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
            return Path(handle.name)
        finally:
            handle.close()

    @staticmethod
    def _atomic_write_bytes(path: Path, payload: bytes) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        handle = tempfile.NamedTemporaryFile(
            prefix=f".{path.name}.",
            dir=path.parent,
            delete=False,
        )
        temp_path = Path(handle.name)
        try:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
            handle.close()
            os.replace(temp_path, path)
        finally:
            try:
                handle.close()
            except Exception:
                pass
            temp_path.unlink(missing_ok=True)

    @classmethod
    def _atomic_write_json(cls, path: Path, value: dict) -> None:
        cls._atomic_write_bytes(
            path,
            json.dumps(
                value,
                ensure_ascii=False,
                indent=2,
                sort_keys=True,
            ).encode("utf-8"),
        )
